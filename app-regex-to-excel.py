from openpyxl import Workbook
import re

'''
FALTA AJUSTES!!!!!!!!!!!!!!!!
'''

def extrair_padroes(content, patterns):
    """Extrai padrões do conteúdo e retorna como lista."""
    return {key: re.findall(pattern, content) for key, pattern in patterns.items()}


def agrupar_pares(lista, step=2):
    """Agrupa itens de uma lista em pares."""
    return [tuple(lista[i:i + step]) for i in range(0, len(lista), step)]


def processar_arquivo(input_file, output_file):
    # Padrões para busca
    patterns = {
        "serial_MONO": r"Número de série:X3BK03\w{4,}",
        "serial_COLOR": r"Número de série:XBJZ02\w{4,}",
        "serial_COLOR_2": r"Número de série: XBJ\w{4,}", 
        "total_count": r"Número total de páginas:\w{3,}",
        "paginas_pb_cor": r"Número total de páginas a .{3,}"
    }

    # Ler o conteúdo do arquivo
    with open(input_file, 'r', encoding='utf-8') as file:
        content = file.read()

    # Extrair padrões
    dados = extrair_padroes(content, patterns)
    # print(dados)

    # Verificar consistência de dados extraídos
    if not dados["serial_MONO"] or not dados["serial_COLOR"]:
        print("Erro: Não foram encontrados números de série suficientes.")
        return

    if len(dados["paginas_pb_cor"]) % 2 != 0:
        print("Erro: Número de páginas P&B e Coloridas não estão agrupados corretamente.")
        return

    # Agrupar contadores P&B e Coloridos
    paginas_pb_cor_pares = agrupar_pares(dados["paginas_pb_cor"])

    # Criar zips
    zip_MONO = list(zip(dados["serial_MONO"], dados["total_count"]))
    zip_COLOR = list(zip(dados["serial_COLOR"], paginas_pb_cor_pares))
    zip_COLOR_2 = list(zip(dados["serial_COLOR_2"], paginas_pb_cor_pares))


    print('Informações extraídas com sucesso!')

    return zip_MONO, zip_COLOR, zip_COLOR_2


# Executar função principal
a, b, c = processar_arquivo('texto-extraido.txt', 'texto-filtrado.txt')

# print(a)
print()
print(b)
print()
# print(c)


wb = Workbook()

ws = wb.active

# Define título para o sheet
ws.title = "Contagem extraída"

ws['A1'] = "Número de série"
ws['B1'] = "Número total de páginas"

for i in a:
    ws.append(i)

for i in b:
    ws.append(i[0])
    ws.append(i[1])

for i in c:
    ws.append(i)

wb.save("sample.xlsx")